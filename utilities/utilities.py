import csv, os, itertools, re, sys, urllib2, datetime, pdb
from urlparse import urlsplit


def gce_webservice( baseUrl, outputLocation, delimiter=None, **kwargs ):

        '''
        The 'baseUrl' is the base url for the webservice.  You are probably using
        the GCE Metabase Web Service API.  This code was written to work with this API as of May
        2011.  The base url is (or is something like):

        baseUrl = 'http://gce-lter.marsci.uga.edu/public/app/location_search.asp?'

        The 'outputLocation' is the location where the generated csv file will
        be saved. It should be something like:

        outputLocation = 'C:\UserFiles\Path\ToSome\Folder'

        The 'delimiter' argument is the character used in the webservice to
        seperate parameters.  This is an optional parameter.  For instance the 
        GCE Metabase web service uses a '&'. For example:

        http://gce-lter.marsci.uga.edu/public/app/location_search.asp?type=sonde&name=GCE10

        However, if the webservice you are calling uses a different 'delimiter'
        then you can specifiy it withe the delimiter argument as:

        delimiter = '$$'

        Or whatever the correct delimiter is.  
     
        
        The '**kwargs' are comma seperated queries supported by the GCE Metabase
        Web Service API.  No keyword arguments will simply return everything.
        The '**kwargs' should be something like this:

        fetchByUrl( baseUrl, outputLocation, type='sonde', name='GCE10', shapefile='yes')

        You probably want the kwarg "shapefile='yes'" because it will generate
        a csv file with headers that have less than 10 characters.  The reason
        that that is important is because ArcGis can not generate attribute
        columns with names longer than 10 characters.  The 'makeShapefile'
        script (that uses this module) truncates a header at 10 characters so it will not blow up in
        ArcGIS.  This paramater tells the web service to make headers with less
        than 10 characters using a controlled vocabulary that we defined (ie is
        legable). Otherwise the script will truncate the headers at 10
        characters, and they might not make sense to the user.

        '''

        # If the delimiter argument is 'None', then use the delimiter for the
        # GCE Metabase API
        if delimiter is None:

            delimiter = '&'

        # if the delimiter is specified as something else, then use that
        # delimiter
        else:

            # Set the delimiter as what is passed in 
            delimiter = delimiter

        # Convert dictionary of kwargs to a list
        dict_to_list = kwargs.items()

        # For item in the list we want to join with a '=' to build url for the web service
        for item in dict_to_list:

            # Grab the index so we can replace the item in the list with the correctly edited item
            index = dict_to_list.index( item )

            # Replace the item in the list with the joined item
            dict_to_list[ index ] = ('=').join( item )

        # Join every item in the list with a '&'.  The '&' is used in the GCE
        # Metabase Web Service to join parameters
	dict_to_list = ( delimiter ).join( dict_to_list )

        # Build the final url we are going to use for the web service
	url = baseUrl + dict_to_list

        print 'Url passed to Web Service: ', url

        # Object that will hold the request
        req = urllib2.Request( url )

        # Object that holds the response
	response = urllib2.urlopen( req )

        # variable that holds the response as a string
	the_page = response.read()

        if response.info().has_key('Content-Disposition'):
        
            # Call the headers so we can name the file according to the filename it is downloaded as
            headers = response.info()

            # Grab the 'Content-Disposition' of the header and split and get the first index to the file name
            filename = headers['Content-Disposition'].split('=')[1]

            # Grab all the files in output location to make sure that there is not
            # already a file named the same as the one we just got.  The GCE
            # webservice does not name the files uniquely, but rather it just gives
            # one generic file name. So, if you are retrieving many files by
            # 'fetchUrl' method then you are going to write over every one, because
            # they all would have the same name
            allFiles_outputLocation = os.listdir( outputLocation )

            # The GCE webservices is doing something weird. It allows the csv file
            # to be opened in Excel, but it is not able to programmaticaly parse
            # them with the csv reader library. So, the code below fixes it
            the_page = the_page.replace('\r','\n')

            # If the filename is NOT the same as the filename we have, then name it
            # as what get retrieved
            if filename not in allFiles_outputLocation:

                filepath = os.path.join( outputLocation, filename)
                file = open( filepath , 'w')
                file.write(the_page)
                file.close()
                print 'File saved to: ', filepath 

            # If there is already a file name what we have retrieved then give a unique name
            else:

                # Split the filename
                filename, extension = filename.split('.')

                # Get a datetime. This will allow us to name it uniquely
                tag = str(datetime.datetime.now())

                # A list that holds all the characters we need to replace.
                # Otherwise it would blow up in ArcGIS
                replace = ['-','.',':',' ']

                # For each item in the replace list...replace them
                for each in replace:
                    tag = tag.replace( each, '')

                # Construct the filename
                filename = filename + '_' + tag + '.' + extension

                file = open( os.path.join( outputLocation, filename), 'w')
                file.write(the_page)
                file.close()
                print 'File saved to: ', os.path.join( outputLocation, filename) 

        else:

            print "File does not have a 'Content-Disposition' key."
            print "May have been redirected. Add code to handle this. ",
            print "See 'http://stackoverflow.com/questions/862173/how-to-download-a-file-using-python-in-a-smarter-way",


def dms_to_dd( pathToInput, pathToOutput, lat_index=None, long_index=None ):

    '''
    Summary: Convert Degrees Minutes and Seconds to Decimal Degrees.  This script
    should be able (although not tested thoroughly) be able to parse the
    Degrees Minutes and Seconds for latitude in one column and Degrees Minutes
    and Seconds in for longitude in another column.  The script parses out the
    degree symbol and all other characters.
    
    'pathToInput' is the path to the csv file you want to work on 

    The pathToInput csv file must be formatted like:

    lattideInFirstColmn, longitudeinSecondColumn 
    31(degree symbol)22'43.56"N, 81(degree symbol)17'24.32"W
    31(degree symbol)22'33.12"N, 81(degree symbol)10'12.21"W

    The first row has to be in this order, and that is latitude and the
    longitude.

    'pathtoOutput' is the path to the FOLDER where you want the csv file saved to

    'lat_index' - header column index that contains the latitude coordinates

    'long_index' - header column index that contains the latitude coordinates

    The output is a csv file with the coordinates coverted to Decimal Degrees.
    It is named the same as the input file, but has 'output_' prepended to it.

    FirstColumnName, SecondColumnName 
    31.41277, 81.29027
    31.41279, 81.29016
    
    '''

    #import csv, os, itertools, re, pdb

    # Open file. Has to be at before the 'header == None' code just below 
    f = open( pathToInput , 'r')

    headers = f.readline().split(',')

    # Folder where you want to write your output files
    outputFilePath = os.path.join( pathToOutput , os.path.basename( pathToInput ))

    # Opening the output file
    outputFile = open( outputFilePath, 'w')
    
    # This code is here in case a list of headers was not passed in.  It strips
    # out any carraige returns that would otherwise blow the script up.  This
    # has to be done because we are reading the first line and not using the
    # csv reader (which would already have parsed out the any problems)
    headers = [each.replace('\n','') for each in headers]

    # Deleting any blank cells that may have been there if the creater of the
    # csv file filled out a cell and then erased.  If that was the case then
    # there would be a blank cell and corresponding list item.
    for header in headers:

        if header == '':

            del headers[ headers.index( header ) ]


    # Using DictReader from csv library and the paramaters we made
    reader = csv.DictReader( f, delimiter = ',', fieldnames = headers )

    # If the lat_index and long_index arguments were passed in, then make the headers list
    # contain just those items
    if lat_index != None and long_index != None:

        headers = [ headers[ lat_index ], headers[ long_index ] ]

    # A cool way to count. Used for counting the rows for print out
    inc = itertools.count(1).next

    # Writing the headers out to file
    outputFile.write('%s,%s\n' % ( headers[0], headers[1] ))

    # Iterating over the csv information
    for row in reader:

        for header in headers:

            # Strip off any spaces before or after the row values
            row[ header ] = row[ header ].strip()

            # RE to find any degree symbols and things would want to replace. 
            pat = r'[^0-9.NnWwSsEe"]+'

            # uses 'findall' to make a list of everything that was found
            findall = re.findall(pat, row[ header ])

            # Will return True if it found anything was found in the regular expression
            if findall:
                    
                    # iterate over everything that was found in the 'findall'
                    for found in findall:

                        # Replace eveyrying that was found in the regex with '_$__' which is used as a place mark for parsing later.
                        row[ header ] = row[ header ].replace( found, '_$_')

                    # Items that need to removed from the dms coordinate
                    to_remove = [ 'W','w', 'E','e','N','n', 'S','s','"']

                    # Items that if found need to have a '-' prepended to, because the dd should be negative
                    make_negative = ['W','w', 'S','s']

                    # Loop through each item in to_remove and if found then loop through the list
                    # make_negative and if found the prepend '-' to row[header] if there is not
                    # already a negative there.
                    for remove in to_remove:
                        if remove in row[ header ]:
                            for negative in make_negative:
                                if negative in row[ header ]:
                                    if '-' not in row[ header ]:
                                        row[ header ] = '-' + row[ header ]

                            # Replace the 'remove' with nothing
                            row[ header ] = row[ header ].replace( remove, '')

                    # Split each at the '_$_' that we put in to replace all the things we didn't want
                    row[ header ] = row[ header ].split('_$_')

            # If only integers or float data types were found
            else:

                print 'File not processed.'


        # Try the following. if it fails the there is only a print statement
        try:

            '''
            print row[headers[0]]
            print row[ headers[0] ] [0]
            print row[ headers[0] ] [1]
            print row[ headers[0] ] [2]
            
            print row[headers[1]]
            print row[ headers[1] ] [0]
            print row[ headers[1] ] [1]
            print row[ headers[1] ] [2]
            '''

            # For the longitude, if there was a 'negative' sign in it, then
            # replace with nothing for the moment.  We have to do this or else
            # the calculation will be wrong. We will multiply it all by '-1' in a little bit
            if '-' in row[ headers [1] ][0]:

                # Replace the negataive
                row[ headers [1] ][0] = row[ headers [1] ][0].replace('-','')
                multiply_by = -1

            else:
                multiply_by = 1

            # Calculate longitude
            DdLon = str(( float( row[ headers[1] ] [0] ) + ( float(row[ headers[1] ] [1]) * 1/60) + ( float(row[ headers[1] ] [2]) * 1/60 * 1/60 ) ) * multiply_by  )

            if '-' in row[ headers [0] ][0]:

                # Replace the negataive
                row[ headers [0] ][0] = row[ headers [0] ][0].replace('-','')
                multiply_by = -1

            elif '-' not in row[ headers [0] ][0]:
                multiply_by = 1

            # Calculate latitude
            DdLat = str(( float( row[ headers[0] ] [0] ) + ( float(row[ headers[0] ] [1]) * 1/60) + ( float(row[ headers[0] ] [2]) * 1/60 * 1/60 ) ) * multiply_by)

            # Write it all out to file
            outputFile.write( '%s, %s\n' % ( DdLat, DdLon ) )

        # If there was something wrong overall then print that out.
        except ValueError:

            print 'error on line: ', inc()

        
    # Close the files
    f.close()
    outputFile.close()

    print 'Successfully converted.'


def filter_csv( inputFilePath, outputFolderPath, headerNamesOfInterest, replace_errors=None ):

    """
    Summary: Filters a csv file based on the 'headerNamesOfInterest' list
    passed in for every file in a directory.  Only the headers in the list and
    their associated row data will be included in 
    in the new csv file.  A new csv file is created with 'output_' prepended to
    it. Need to change script to operate only on one file.

    'inputFolderPath' - The path to the folder where the files are located. for
    example:

         inputFolderPath = r'C:\UserFiles\path\to\folder'


    'outputFolderPath' - The path to the folder where the files are located. for
    example:

         outputFolderPath = r'C:\UserFiles\path\to\folder'

    'headerNamesOfInterest' - A list containing header names from the csv file.
    All header names included in this list must be spelled exactly the way they
    are in the csv file.

    'replace_errors' - optional. A string that will be used to replace characters that are stopping
    the csv from being processed.

    """

    # Open each file while making the path so the file can be opened
    f = open( inputFilePath, 'r')

    # Opening the output file
    outputFile = open( os.path.join( outputFolderPath, 'output_' + os.path.splitext( os.path.basename( inputFilePath ))[0] + '.csv'), 'w')
        
    # Grabbing the fields (the first line of each file), and splitting them
    # making a list
    fields = f.readline().split(',')

    # This a list comprehension it does the same thing as the code that is
    # commented out, but is is shorter and quicker.  It is a list that holds 
    # the headers after we cleaned them up.  I had to clean them
    # up because some of the headers had '\n' included and an error was thrown
    # meaning it wasn't parsable. This gets rid of the problems.
    headers = [ each.replace( '\n' , '' ) for each in fields ]

    # Using DictReader from csv library and the paramaters we made
    reader = csv.DictReader( f, delimiter = ',', fieldnames = headers )

    # Loops through each item in 'headerNamesOfInterest'. Enumerate lets you
    # grab the index of each item. This allows you to dynamically build the csv
    # headers (ie there can be 1 to a lot of items in this list and it will be
    # created)
    for index, each in enumerate( headerNamesOfInterest ):

        # Write the first header item to the file
        outputFile.write( '%s,' % headerNamesOfInterest[ index ] )

    # When you get to the last item write a '\n'
    outputFile.write('\n')

    # Counter used for printing to see what is going on
    counter = 0

    # For each row in reader object
    for row in reader:
        
        # Try the below
        try:

            # Loop through each row in csv file that is apart of the headersNameOfInterest. If any
            # character in the list of "blowing_it_up" is found then replace it
            for index, value in enumerate( headerNamesOfInterest ):

                # These characters do not allow the csv file to be proceesed 
                blowing_it_up = [ ',', '\n', "'", '"' ]

                for blowup in blowing_it_up:

                    if blowup in row[ headerNamesOfInterest[ index ] ]:

                        print 'A " %s " is not allowing the csv to be processed so it was replaced with a " ".' % blowup

                        # If 'replace_errors_with' was passed in then use custom replace
                        if replace_errors != None:

                            replace_error_with = replace_errors

                        # if it was not passed in then replace with no space
                        else:

                            replace_error_with = ''

                        # replace any characters that were blowing it up
                        row[ headerNamesOfInterest[ index ] ] =  row[ headerNamesOfInterest[ index ] ].replace( blowup, replace_error_with )

                # Add one to the counter
                counter += 1

                # Write the item to file
                outputFile.write( '%s,' % ( row[ headerNamesOfInterest[index] ]))

            # When you get to the last item write a '\n'
            outputFile.write('\n')

        except:

            print 'File was not able to be processed'
                
    # Close the files
    f.close()
    outputFile.close()

    print 'Filter succesfully processed'


def excel_to_csv( pathToTopFolder, output_path ):
    """
    Summary: Recursively walk through a directory and convert xls and xlsx files to csv files

    'pathToTopFolder' - Path to the root of the directory that you want to start recursively walking through

    'output_path' - Path to directory where you want the csv files saved to
    """

    # Recursively walk through each directory and file
    for (path, dirs, files) in os.walk( pathToTopFolder ):

        # Loop through files
        for File in files:

            # Get root name and file extension
            root, ext = os.path.splitext( os.path.basename( File ))
           
            if ext == '.xls' :

                # Make input_file_path
                inputFilePath = os.path.join( path, File )

                # Invoke function to convert to xls files to csv
                xls_to_csv( inputFilePath, output_path )

            if ext == '.xlsx':
                
                inputFilePath = os.path.join( path, File )  

                xlsx_to_csv( inputFilePath, output_path )                       
               

    print 'Files converted to csv'


def xls_to_csv( inputFilePath, output_path ):

    """
    Summary: Convert xls files to csv files

    'pathToTopFolder' - Path to the file that you want to start recursively walking through

    'output_path' - Path to directory where you want the csv files saved to
    """

    import xlrd

    os.chdir( os.path.dirname( inputFilePath) )

    root, ext = os.path.splitext( inputFilePath )
    newFile = root + '.csv'
    
    wb = xlrd.open_workbook( os.path.basename( inputFilePath) )
    wb.sheet_names()
    sh = wb.sheet_by_index(0)

    count = 1

    csvFile = open( os.path.join( output_path , os.path.basename( newFile) ) , 'w' )

    for rownum in range(sh.nrows):

        data = sh.row_values( rownum )

        if count == 1:

            for each in data:

                csvFile.write('%s,' % each)

            csvFile.write('\n')

        else:

            for each in data:

                csvFile.write('%s,' % each)

            csvFile.write('\n')

        count += 1

    csvFile.close()

    print 'xls file converted to csv'


def xlsx_to_csv( inputFilePath, output_path ):

    """
    Summary: Convert xlsx files to csv files

    'pathToTopFolder' - Path to the file that you want to start recursively walking through

    'output_path' - Path to directory where you want the csv files saved to
    """

    from openpyxl.reader.excel import load_workbook
   
    os.chdir( os.path.dirname( inputFilePath) )

    root, ext = os.path.splitext( inputFilePath )
    newFile = root + '.csv'

    # Open csv file
    csvFile = open( os.path.join( output_path , os.path.basename( newFile) ) ,'w')

    # Open xlsx file
    wb = load_workbook(filename = os.path.basename( inputFilePath), use_iterators = True)

    # Get the first sheet
    ws = wb.get_sheet_by_name(name = wb.get_sheet_names()[0] )

    i = 0

    firstRow = []

    for row in ws.iter_rows():

        if i == 0:

            for cell in row:

                firstRow.append( cell.internal_value )
                try:

                    csvFile.write('%s,' % cell.internal_value.replace('\n',''))

                except:

                    csvFile.write('%s,' % cell.internal_value)


            csvFile.write('\n')

        else:

            for cell in row:

                csvFile.write('%s,' % cell.internal_value)

            csvFile.write('\n')

        i += 1

    csvFile.close()

    print 'xlsx file converted to csv'




